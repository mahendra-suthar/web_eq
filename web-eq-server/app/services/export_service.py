from __future__ import annotations

from io import BytesIO
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

MAX_EXPORT_ROWS = 10_000


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def build_xlsx(columns: list[str], rows: list[list[Any]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    ws.append(columns)
    header_fill = PatternFill("solid", fgColor="00695C")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row in rows:
        ws.append([_safe_str(v) for v in row])

    # auto-width capped at 40 characters
    for col in ws.columns:
        col_width = max((len(_safe_str(cell.value)) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(col_width + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(title: str, columns: list[str], rows: list[list[Any]]) -> BytesIO:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=30,
        bottomMargin=20,
        leftMargin=20,
        rightMargin=20,
    )
    styles = getSampleStyleSheet()
    elements: list[Any] = [Paragraph(title, styles["Title"])]

    data = [columns] + [[_safe_str(v) for v in row] for row in rows]
    col_count = len(columns)
    page_width = landscape(A4)[0] - 40  # usable width after margins
    col_width = page_width / col_count

    t = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#00695C")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F8F6")]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf
